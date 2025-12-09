import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image
import numpy as np

# --- CONFIGURACIÓN GENERAL DE LA PÁGINA ---
st.set_page_config(
    page_title="MarketMaster",
    page_icon="🚀",
    layout="wide"
)

# --- LÓGICA PARA MERCADO LIBRE MEDELLÍN ---
def pagina_meli_medellin():
    """
    Contiene toda la lógica y la interfaz para procesar archivos de Mercado Libre Medellín.
    """
    st.markdown("### 🛍️ Mercado Libre - Medellín")
    
    column_names = [
        'Número de publicación', 'Número de producto', 'Número de variante', 'SKU', 'Título', 'Variantes',
        'Cantidad', 'Precio', 'Moneda'
    ]

    uploaded_file_meli = st.file_uploader("📤 Cargar archivo Excel de Mercado Libre", type=['xlsx'], key="meli_med_excel")
    uploaded_file_erp = st.file_uploader("🧾 Cargar archivo CSV de ERP", type=['csv'], key="meli_med_erp")

    if uploaded_file_meli and uploaded_file_erp:
        if st.button('🔄 Procesar MELI Medellín', key="meli_med_process"):
            with st.spinner('Procesando archivos...'):
                try:
                    data_MELI = pd.read_excel(uploaded_file_meli, header=None, skiprows=5, names=column_names, sheet_name="Publicaciones")
                    data_ERP = pd.read_csv(uploaded_file_erp, delimiter=';', encoding='latin1')

                    data_ERP = data_ERP[data_ERP['Codpro'].notna() & ~(data_ERP['Codpro'].isin(['', ' ']) | (data_ERP['Codpro'].str.contains('\x1a', na=False)))]
                    data_ERP = data_ERP[["Codpro", "Nompro", "Valuni", "us05"]]
                    data_ERP['us05'] = data_ERP['us05'].fillna(0)
                    data_ERP["Inventario_Medellin"] = data_ERP["us05"]
                    data_ERP = data_ERP.drop(["us05"], axis=1)
                    data_ERP.rename(columns={'Codpro': 'SKU'}, inplace=True)

                    data_MELI['SKU'] = data_MELI['SKU'].astype(str)
                    data_ERP['SKU'] = data_ERP['SKU'].astype(str)
                    data_MELI['SKU'] = data_MELI['SKU'].replace('nan', np.nan)
                    data_ERP['SKU'] = data_ERP['SKU'].replace('nan', np.nan)

                    merged_data = pd.merge(data_MELI, data_ERP, on='SKU', how='left')
                    merged_data['Original_Price'] = merged_data['Precio']
                    merged_data['original_order'] = merged_data.index

                    grouped = merged_data.groupby('Número de publicación')
                    processed_groups = []
                    for name, group in grouped:
                        if group.shape[0] == 1:
                            group.loc[:, "Cantidad"] = group["Inventario_Medellin"]
                            group.loc[:, "Precio"] = group["Valuni"]
                        elif group.shape[0] > 1:
                            group.loc[group.SKU.notna(), "Cantidad"] = group.loc[group.SKU.notna(), "Inventario_Medellin"]
                            max_price = group.loc[group.SKU.notna(), "Valuni"].max()
                            group.loc[group.SKU.isna(), "Precio"] = max_price
                        processed_groups.append(group)

                    final_df = pd.concat(processed_groups)
                    final_df['Precio'] = final_df['Precio'].fillna(final_df['Original_Price'])
                    final_df = final_df.sort_values('original_order')
                    final_df['Número de variante'] = final_df['Número de variante'].apply(lambda x: str(int(x)) if pd.notna(x) else None)
                    final_df = final_df.drop(['Original_Price', 'Nompro', 'Valuni', 'Inventario_Medellin', 'original_order'], axis=1)

                    wb = load_workbook(uploaded_file_meli)
                    ws = wb['Publicaciones']
                    
                    for r_idx, row_data in final_df.iterrows():
                        for c_idx, value in enumerate(row_data, start=1):
                            ws.cell(row=r_idx + 6, column=c_idx, value=value)

                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.success("✅ ¡Archivo de MELI Medellín procesado!")
                    st.dataframe(final_df.head())
                    st.download_button(label="⬇️ Descargar MELI Medellín modificado",
                                      data=output,
                                      file_name="MELI_Medellin_ACTUALIZADO.xlsx",
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")

# --- LÓGICA PARA MERCADO LIBRE BOGOTÁ ---
def pagina_meli_bogota():
    st.markdown("### 🛒 Mercado Libre - Bogotá")
    column_names = [
        'Número de publicación', 'Número de variante', 'SKU', 'Título', 'Variantes',
        'Cantidad (Obligatorio)', 'Canal de venta', 'Precio', 'Mercado Shops',
        'Vincular precio con Mercado Libre', 'Moneda'
    ]

    uploaded_file_meli = st.file_uploader("📤 Cargar archivo Excel de Mercado Libre", type=['xlsx'], key="meli_bog_excel")
    uploaded_file_erp = st.file_uploader("🧾 Cargar archivo CSV de ERP", type=['csv'], key="meli_bog_erp")

    if uploaded_file_meli and uploaded_file_erp:
        if st.button('🔄 Procesar MELI Bogotá', key="meli_bog_process"):
            with st.spinner('Procesando archivos...'):
                try:
                    data_MELI = pd.read_excel(uploaded_file_meli, header=None, skiprows=6, names=column_names, sheet_name="Publicaciones")
                    data_ERP = pd.read_csv(uploaded_file_erp, delimiter=';', encoding='latin1')

                    data_ERP = data_ERP[data_ERP['Codpro'].notna() & ~(data_ERP['Codpro'].isin(['', ' ']) | (data_ERP['Codpro'].str.contains('\x1a', na=False)))]
                    data_ERP = data_ERP[["Codpro", "Nompro", "Valuni", "us01", "us02"]]
                    data_ERP['us01'] = data_ERP['us01'].fillna(0)
                    data_ERP['us02'] = data_ERP['us02'].fillna(0)
                    data_ERP["Inventario_Bogota"] = data_ERP["us01"] + data_ERP["us02"]
                    data_ERP = data_ERP.drop(["us01", "us02"], axis=1)
                    data_ERP.rename(columns={'Codpro': 'SKU'}, inplace=True)

                    data_MELI['SKU'] = data_MELI['SKU'].astype(str)
                    data_ERP['SKU'] = data_ERP['SKU'].astype(str)
                    data_MELI['SKU'] = data_MELI['SKU'].replace('nan', np.nan)
                    data_ERP['SKU'] = data_ERP['SKU'].replace('nan', np.nan)

                    merged_data = pd.merge(data_MELI, data_ERP, on='SKU', how='left')
                    merged_data['Original_Price'] = merged_data['Precio']
                    merged_data['original_order'] = merged_data.index

                    grouped = merged_data.groupby('Número de publicación')
                    processed_groups = []
                    for name, group in grouped:
                        if group.shape[0] == 1:
                            group.loc[:, "Cantidad (Obligatorio)"] = group["Inventario_Bogota"]
                            group.loc[:, "Precio"] = group["Valuni"]
                        elif group.shape[0] > 1:
                            group.loc[group.SKU.notna(), "Cantidad (Obligatorio)"] = group.loc[group.SKU.notna(), "Inventario_Bogota"]
                            max_price = group.loc[group.SKU.notna(), "Valuni"].max()
                            group.loc[group.SKU.isna(), "Precio"] = max_price
                        processed_groups.append(group)

                    final_df = pd.concat(processed_groups)
                    final_df['Precio'] = final_df['Precio'].fillna(final_df['Original_Price'])
                    final_df = final_df.sort_values('original_order')
                    final_df['Número de variante'] = final_df['Número de variante'].apply(lambda x: str(int(x)) if pd.notna(x) else None)
                    final_df = final_df.drop(['Nompro', 'Valuni', 'Inventario_Bogota', 'original_order', 'Original_Price'], axis=1)

                    wb = load_workbook(uploaded_file_meli)
                    ws = wb['Publicaciones']
                    for r_idx, row_data in final_df.iterrows():
                        for c_idx, value in enumerate(row_data, start=1):
                            ws.cell(row=r_idx + 7, column=c_idx, value=value)
                    
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.success("✅ ¡Archivo de MELI Bogotá procesado!")
                    st.dataframe(final_df.head())
                    st.download_button(label="⬇️ Descargar MELI Bogotá modificado",
                                      data=output,
                                      file_name="MELI_Bogota_ACTUALIZADO.xlsx",
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")

# --- LÓGICA PARA FALABELLA ---
def pagina_falabella():
    st.markdown("### 🧩 Falabella")
    
    column_names_price = ['SellerSku', 'ShopSku', 'PriceFalabella', 'SalePriceFalabella', 'SaleStartDateFalabella', 'SaleEndDateFalabella', 'Name']
    column_names_inventory = ['SellerSku', 'ShopSku', 'QuantityFalabella', 'Name']

    uploaded_price = st.file_uploader("📦 Cargar archivo de precios (Excel)", type=['xlsx'], key="fala_price")
    uploaded_inventory = st.file_uploader("📊 Cargar archivo de inventario (CSV)", type=['csv'], key="fala_inv")
    uploaded_erp = st.file_uploader("🧾 Cargar archivo ERP (CSV)", type=['csv'], key="fala_erp")

    if uploaded_price and uploaded_inventory and uploaded_erp:
        if st.button("🔄 Procesar Falabella", key="fala_process"):
            with st.spinner('Procesando archivos...'):
                try:
                    data_price = pd.read_excel(uploaded_price, header=None, skiprows=1, names=column_names_price)
                    data_inventory = pd.read_csv(uploaded_inventory, header=None, skiprows=1, names=column_names_inventory, sep=';', encoding='utf-8')
                    data_erp = pd.read_csv(uploaded_erp, delimiter=';', encoding='latin1')

                    data_erp = data_erp[data_erp['Codpro'].notna() & ~(data_erp['Codpro'].isin(['', ' ']) | data_erp['Codpro'].str.contains('\x1a', na=False))]
                    data_erp = data_erp[['Codpro', 'Nompro', 'Valuni', 'us01', 'us02']]
                    data_erp['us01'] = data_erp['us01'].fillna(0)
                    data_erp['us02'] = data_erp['us02'].fillna(0)
                    data_erp['Inventario_Bogota'] = data_erp['us01'] + data_erp['us02']
                    data_erp.drop(['us01', 'us02'], axis=1, inplace=True)
                    data_erp.rename(columns={'Codpro': 'sku'}, inplace=True)

                    for df in [data_price, data_inventory]:
                        df.rename(columns={'SellerSku': 'sku'}, inplace=True)
                    for df in [data_price, data_inventory, data_erp]:
                        df['sku'] = df['sku'].astype(str).str.strip()
                    
                    data_price['ShopSku'] = data_price['ShopSku'].astype(str).str.replace('.0', '', regex=False)
                    data_inventory['ShopSku'] = data_inventory['ShopSku'].astype(str).str.replace('.0', '', regex=False)

                    # Procesar precios
                    st.info("Procesando archivo de precios...")
                    merged_price = pd.merge(data_price, data_erp[['sku', 'Valuni']], on='sku', how='left')
                    merged_price['PriceFalabella'] = merged_price['Valuni']
                    merged_price.drop(columns=['Valuni'], inplace=True)
                    
                    wb_price = load_workbook(uploaded_price)
                    ws_price = wb_price.active
                    for i, row in merged_price.iterrows():
                        for j, value in enumerate(row):
                            ws_price.cell(row=i+2, column=j+1, value=value)
                    
                    buffer_price = BytesIO()
                    wb_price.save(buffer_price)
                    buffer_price.seek(0)
                    
                    # Procesar inventario
                    st.info("Procesando archivo de inventario...")
                    merged_inventory = pd.merge(data_inventory, data_erp[['sku', 'Inventario_Bogota']], on='sku', how='left')
                    merged_inventory['QuantityFalabella'] = merged_inventory['Inventario_Bogota'].fillna(0).astype('int')
                    merged_inventory.drop(columns=['Inventario_Bogota'], inplace=True)
                    merged_inventory.rename(columns={'sku': 'SellerSku'}, inplace=True)
                    
                    csv_data = merged_inventory.to_csv(index=False, sep=';', encoding='utf-8-sig')

                    st.success("✅ ¡Archivos de Falabella procesados!")
                    
                    st.download_button("⬇️ Descargar precios modificados", buffer_price, "Precios_Falabella_Modificado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.download_button("⬇️ Descargar inventario modificado", csv_data, "Inventario_Falabella_Modificado.csv", mime="text/csv")
                    
                    st.markdown("#### Vista Previa Inventario")
                    st.dataframe(merged_inventory.head())

                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")

# --- LÓGICA PARA RAPPI (BOGOTÁ STORES) ---
def pagina_rappi_av19_blv_cll74():
    st.markdown("### 🛵 Rappi - Av.19, Blv y Cll 74")
    column_names = ['vacia_borrar', 'ID', 'ID de tienda', 'Nombre de tienda', 'ID del producto', 'EAN', 'SKU' ,'Nombre del producto', 'Descripción', 'Presentación', 'Precio', 'Descuento', 'Disponibilidad']
    mapeo_tienda_us = { 900243006: 'us01', 900243075: 'us02', 900246112: 'us03' }

    uploaded_file_rappi = st.file_uploader("📤 Cargar archivo Excel de Rappi", type=['xlsx'], key="rappi_bog_excel")
    uploaded_file_erp = st.file_uploader("🧾 Cargar archivo CSV de ERP", type=['csv'], key="rappi_bog_erp")

    if uploaded_file_rappi and uploaded_file_erp:
        if st.button('🔄 Procesar Rappi (Av.19, Blv y Cll 74)', key="rappi_bog_process"):
            with st.spinner('Procesando archivos...'):
                try:
                    data_RAPPI = pd.read_excel(uploaded_file_rappi, header=None, skiprows=5, names=column_names, sheet_name="Productos")
                    data_ERP = pd.read_csv(uploaded_file_erp, delimiter=';', encoding='latin1')

                    data_ERP = data_ERP[data_ERP['Codpro'].notna() & ~(data_ERP['Codpro'].isin(['', ' ']) | (data_ERP['Codpro'].str.contains('\x1a', na=False)))]
                    data_ERP = data_ERP[["Codpro", "Nompro", "Valuni", "us01", "us02", "us03"]]
                    data_ERP.rename(columns={'Codpro': 'SKU'}, inplace=True)

                    data_RAPPI['SKU'] = data_RAPPI['SKU'].astype(str).str.replace('jugandoyeducandoco_', '')
                    data_ERP['SKU'] = data_ERP['SKU'].astype(str)
                    data_RAPPI['tienda_us'] = data_RAPPI['ID de tienda'].map(mapeo_tienda_us)

                    def obtener_inventario(row, df_erp):
                        col_inv = row['tienda_us']
                        sku = row['SKU']
                        if pd.notna(col_inv) and pd.notna(sku):
                            inventario = df_erp.loc[df_erp['SKU'] == sku, col_inv]
                            return int(inventario.iloc[0]) if not inventario.empty and pd.notna(inventario.iloc[0]) else 0
                        return 0

                    data_RAPPI['Inventario'] = data_RAPPI.apply(obtener_inventario, df_erp=data_ERP, axis=1)
                    data_RAPPI['Disponibilidad_correcta'] = np.where(data_RAPPI['Inventario'] > 0, 'SI', 'NO')

                    merged_data = pd.merge(data_RAPPI, data_ERP, on='SKU', how='left')
                    merged_data['precio_correcto'] = merged_data['Valuni']

                    columnas_deseadas = ['vacia_borrar', 'ID', 'ID de tienda', 'Nombre de tienda', 'ID del producto', 'EAN', 'SKU', 'Nombre del producto', 'Descripción', 'Presentación', 'precio_correcto', 'Descuento', 'Disponibilidad_correcta']
                    nuevo_df_rappi = merged_data[columnas_deseadas].copy()
                    nuevo_df_rappi['SKU'] = "jugandoyeducandoco_" + nuevo_df_rappi['SKU'].astype(str)

                    wb = load_workbook(uploaded_file_rappi)
                    ws = wb['Productos']
                    for index, row in nuevo_df_rappi.iterrows():
                        for col, value in enumerate(row, start=1):
                           ws.cell(row=index + 6, column=col, value=value)

                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.success("✅ ¡Archivo de Rappi (Av.19, Blv y Cll 74) procesado!")
                    st.dataframe(nuevo_df_rappi.head())
                    st.download_button(label="⬇️ Descargar Rappi (Av.19, Blv y Cll 74) modificado",
                                       data=output,
                                       file_name="RAPPI_Av19_Blv_Cll74_ACTUALIZADO.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")

# --- LÓGICA PARA RAPPI (MEDELLÍN STORES) ---
def pagina_rappi_bvista_oviedo():
    st.markdown("### 🛵 Rappi - Bvista y Oviedo")
    column_names = ['vacia_borrar', 'ID', 'ID de tienda', 'Nombre de tienda', 'ID del producto', 'EAN', 'SKU' ,'Nombre del producto', 'Descripción', 'Presentación', 'Precio', 'Descuento', 'Disponibilidad']
    mapeo_tienda_us = { 900243002: 'us04', 900418701: 'us05' }

    uploaded_file_rappi = st.file_uploader("📤 Cargar archivo Excel de Rappi", type=['xlsx'], key="rappi_med_excel")
    uploaded_file_erp = st.file_uploader("🧾 Cargar archivo CSV de ERP", type=['csv'], key="rappi_med_erp")

    if uploaded_file_rappi and uploaded_file_erp:
        if st.button('🔄 Procesar Rappi (Bvista y Oviedo)', key="rappi_med_process"):
            with st.spinner('Procesando archivos...'):
                try:
                    data_RAPPI = pd.read_excel(uploaded_file_rappi, header=None, skiprows=5, names=column_names, sheet_name="Productos")
                    data_ERP = pd.read_csv(uploaded_file_erp, delimiter=';', encoding='latin1')

                    data_ERP = data_ERP[data_ERP['Codpro'].notna() & ~(data_ERP['Codpro'].isin(['', ' ']) | (data_ERP['Codpro'].str.contains('\x1a', na=False)))]
                    data_ERP = data_ERP[["Codpro", "Nompro", "Valuni", "us04", "us05"]]
                    data_ERP.rename(columns={'Codpro': 'SKU'}, inplace=True)

                    data_RAPPI['SKU'] = data_RAPPI['SKU'].astype(str).str.replace('jugandoyeducandoco_', '')
                    data_ERP['SKU'] = data_ERP['SKU'].astype(str)
                    data_RAPPI['tienda_us'] = data_RAPPI['ID de tienda'].map(mapeo_tienda_us)

                    def obtener_inventario(row, df_erp):
                        col_inv = row['tienda_us']
                        sku = row['SKU']
                        if pd.notna(col_inv) and pd.notna(sku):
                            inventario = df_erp.loc[df_erp['SKU'] == sku, col_inv]
                            return int(inventario.iloc[0]) if not inventario.empty and pd.notna(inventario.iloc[0]) else 0
                        return 0

                    data_RAPPI['Inventario'] = data_RAPPI.apply(obtener_inventario, df_erp=data_ERP, axis=1)
                    data_RAPPI['Disponibilidad_correcta'] = np.where(data_RAPPI['Inventario'] > 0, 'SI', 'NO')

                    merged_data = pd.merge(data_RAPPI, data_ERP, on='SKU', how='left')
                    merged_data['precio_correcto'] = merged_data['Valuni']

                    columnas_deseadas = ['vacia_borrar', 'ID', 'ID de tienda', 'Nombre de tienda', 'ID del producto', 'EAN', 'SKU', 'Nombre del producto', 'Descripción', 'Presentación', 'precio_correcto', 'Descuento', 'Disponibilidad_correcta']
                    nuevo_df_rappi = merged_data[columnas_deseadas].copy()
                    nuevo_df_rappi['SKU'] = "jugandoyeducandoco_" + nuevo_df_rappi['SKU'].astype(str)

                    wb = load_workbook(uploaded_file_rappi)
                    ws = wb['Productos']
                    for index, row in nuevo_df_rappi.iterrows():
                        for col, value in enumerate(row, start=1):
                           ws.cell(row=index + 6, column=col, value=value)

                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.success("✅ ¡Archivo de Rappi (Bvista y Oviedo) procesado!")
                    st.dataframe(nuevo_df_rappi.head())
                    st.download_button(label="⬇️ Descargar Rappi (Bvista y Oviedo) modificado",
                                       data=output,
                                       file_name="RAPPI_Bvista_Oviedo_ACTUALIZADO.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")

# --- LÓGICA PARA WIX ---
def pagina_wix():
    st.markdown("## 🌐 Wix")
    column_names = [
        'handleId', 'fieldType', 'name', 'description', 'productImageUrl', 'collection', 'sku', 'ribbon', 
        'price', 'surcharge', 'visible', 'discountMode', 'discountValue', 'inventory', 'weight', 'cost',
        'productOptionName1', 'productOptionType1', 'productOptionDescription1', 'productOptionName2', 'productOptionType2', 'productOptionDescription2',
        'productOptionName3', 'productOptionType3', 'productOptionDescription3', 'productOptionName4', 'productOptionType4', 'productOptionDescription4',
        'productOptionName5', 'productOptionType5', 'productOptionDescription5', 'productOptionName6', 'productOptionType6', 'productOptionDescription6',
        'additionalInfoTitle1', 'additionalInfoDescription1', 'additionalInfoTitle2', 'additionalInfoDescription2',
        'additionalInfoTitle3', 'additionalInfoDescription3', 'additionalInfoTitle4', 'additionalInfoDescription4',
        'additionalInfoTitle5', 'additionalInfoDescription5', 'additionalInfoTitle6', 'additionalInfoDescription6',
        'customTextField1', 'customTextCharLimit1', 'customTextMandatory1', 'customTextField2', 'customTextCharLimit2', 'customTextMandatory2', 'brand'
    ]

    uploaded_file_wix = st.file_uploader("📤 Cargar archivo CSV de Wix", type=['csv'], key="wix_csv")
    uploaded_file_erp = st.file_uploader("🧾 Cargar archivo CSV de ERP", type=['csv'], key="wix_erp")

    if uploaded_file_wix and uploaded_file_erp:
        if st.button('🔄 Procesar Wix', key="wix_process"):
            with st.spinner('Procesando archivos...'):
                try:
                    data_wix = pd.read_csv(uploaded_file_wix, header=0, dtype={'sku': str})
                    # Renombrar columnas después de cargar
                    data_wix.columns = column_names
                    
                    data_ERP = pd.read_csv(uploaded_file_erp, delimiter=';', encoding='latin1')
                    
                    data_ERP = data_ERP[data_ERP['Codpro'].notna() & ~(data_ERP['Codpro'].isin(['', ' ']) | (data_ERP['Codpro'].str.contains('\x1a', na=False)))]
                    data_ERP = data_ERP[["Codpro", "Nompro", "Valuni", "us01", "us02"]]
                    data_ERP['us01'] = data_ERP['us01'].fillna(0)
                    data_ERP['us02'] = data_ERP['us02'].fillna(0)
                    data_ERP["Inventario_Bogota"] = data_ERP["us01"] + data_ERP["us02"]
                    data_ERP.drop(["us01", "us02"], axis=1, inplace=True)
                    data_ERP.rename(columns={'Codpro': 'sku'}, inplace=True)
                    data_ERP['sku'] = data_ERP['sku'].astype(str)

                    merged_data = pd.merge(data_wix, data_ERP, on='sku', how='left')
                    merged_data['Valuni'].fillna(0, inplace=True)
                    merged_data['Inventario_Bogota'].fillna(0, inplace=True)
                    merged_data['inventory'] = merged_data['Inventario_Bogota']
                    merged_data['price'] = merged_data['Valuni']
                    merged_data = merged_data.drop(["Nompro", "Valuni", "Inventario_Bogota"], axis=1)

                    merged_data['visible'] = np.where(merged_data['inventory'] > 0, "TRUE", "FALSE")
                    
                    st.success("✅ ¡Archivo de Wix procesado!")
                    st.dataframe(merged_data.head())
                    
                    num_rows = merged_data.shape[0]
                    max_rows_per_file = 4000
                    num_files = (num_rows // max_rows_per_file) + (1 if num_rows % max_rows_per_file > 0 else 0)
                    
                    st.info(f"El archivo se dividirá en {num_files} parte(s).")
                                    
                    for i in range(num_files):
                        part = merged_data.iloc[i * max_rows_per_file : (i + 1) * max_rows_per_file]
                        output = part.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label=f"⬇️ Descargar Parte {i+1}",
                            data=output,
                            file_name=f"Wix_modificado_parte_{i+1}.csv",
                            mime="text/csv",
                            key=f"wix_download_{i}"
                        )

                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")

# --- LÓGICA PARA WIX VIA API ---
def pagina_wix_api():
    st.markdown("## 🌐 Wix (Actualización Automática vía API)")
    
    st.info("💡 Esta opción sincroniza automáticamente tus productos de Wix con el ERP sin necesidad de cargar archivos CSV de Wix.")
    
    # ============================================================================
    # CARGAR API KEY DESDE SECRETS (OBLIGATORIO)
    # ============================================================================
    try:
        wix_api_key = st.secrets["wix_api"]["api_key"]
        wix_site_id_default = st.secrets["wix_api"]["site_id"]
        st.success("🔒 API Key cargada desde configuración segura")
    except Exception as e:
        st.error("❌ **API Key no configurada**")
        st.warning("⚠️ La API Key debe estar configurada en Streamlit Secrets por seguridad.")
        st.info("📧 Contacta al administrador para configurar las credenciales.")
        st.stop()
    
    # ============================================================================
    # SITE ID - EDITABLE EN UI
    # ============================================================================
    with st.expander("⚙️ Configuración del Sitio", expanded=True):
        st.info("💡 Puedes cambiar el Site ID para probar en diferentes sitios de Wix.")
        wix_site_id = st.text_input(
            "Site ID de Wix",
            value=wix_site_id_default,
            help="El ID de tu sitio Wix (editable para cambiar entre sitios de pruebas y producción)",
            key="wix_site_id_input"
        )
        
        # Mostrar info del sitio
        if wix_site_id == "c9fb6114-70ad-4de2-ba0c-a1bb98e25883":
            st.success("🧪 Sitio de pruebas/backup seleccionado")
        elif wix_site_id == "a290c1b4-e593-4126-ae4e-675bd07c1a42":
            st.warning("🏭 Sitio de producción seleccionado - ¡Ten cuidado!")
        else:
            st.info("🌐 Sitio personalizado")
    
    # Upload del archivo ERP
    uploaded_file_erp = st.file_uploader("🧾 Cargar archivo CSV de ERP", type=['csv'], key="wix_api_erp")
    
    if wix_api_key and wix_site_id and uploaded_file_erp:
        
        # Configurar headers
        headers = {
            'Authorization': wix_api_key,
            'wix-site-id': wix_site_id,
            'Content-Type': 'application/json'
        }
        
        base_url = "https://www.wixapis.com/stores"
        
        # ====================================================================
        # BOTONES DE ANÁLISIS Y SINCRONIZACIÓN
        # ====================================================================
        st.markdown("---")
        col1, col2 = st.columns([1, 1])
        
        with col1:
            analyze_button = st.button('📊 Analizar Cambios', key="wix_api_analyze", use_container_width=True)
        
        with col2:
            sync_button = st.button('🚀 Sincronizar con Wix', key="wix_api_sync", type="primary", use_container_width=True)
        
        # ====================================================================
        # ANÁLISIS PREVIO (sin modificar nada)
        # ====================================================================
        if analyze_button:
            with st.spinner('📊 Analizando productos y calculando cambios...'):
                try:
                    # ============================================================
                    # PASO 1: CARGAR DATOS DEL ERP
                    # ============================================================
                    data_ERP = pd.read_csv(uploaded_file_erp, delimiter=';', encoding='latin1')
                    data_ERP = data_ERP[data_ERP['Codpro'].notna() & ~(data_ERP['Codpro'].isin(['', ' ']) | (data_ERP['Codpro'].str.contains('\x1a', na=False)))]
                    data_ERP = data_ERP[["Codpro", "Nompro", "Valuni", "us01", "us02"]]
                    data_ERP['us01'] = data_ERP['us01'].fillna(0)
                    data_ERP['us02'] = data_ERP['us02'].fillna(0)
                    data_ERP["Inventario_Bogota"] = data_ERP["us01"] + data_ERP["us02"]
                    data_ERP = data_ERP[["Codpro", "Valuni", "Inventario_Bogota"]]
                    data_ERP.rename(columns={'Codpro': 'SKU', 'Valuni': 'Precio', 'Inventario_Bogota': 'Inventario'}, inplace=True)
                    data_ERP['SKU'] = data_ERP['SKU'].astype(str)
                    
                    # ============================================================
                    # PASO 2: DESCARGAR PRODUCTOS DE WIX
                    # ============================================================
                    all_products = []
                    offset = 0
                    limit = 100
                    url = f"{base_url}/v1/products/query"
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Primera llamada para obtener el total
                    payload = {
                        "includeHiddenProducts": True,
                        "query": {"paging": {"limit": limit, "offset": 0}}
                    }
                    
                    response = requests.post(url, headers=headers, json=payload, timeout=30)
                    
                    if response.status_code != 200:
                        st.error(f"❌ Error al conectar con Wix: {response.status_code}")
                        st.code(response.text)
                        st.stop()
                    
                    data = response.json()
                    total_products = data.get("totalResults", 0)
                    all_products.extend(data.get("products", []))
                    
                    status_text.text(f"Descargados: {len(all_products)}/{total_products}")
                    progress_bar.progress(min(len(all_products) / total_products, 1.0))
                    
                    # Continuar descargando el resto
                    offset = limit
                    while len(all_products) < total_products:
                        payload["query"]["paging"]["offset"] = offset
                        
                        response = requests.post(url, headers=headers, json=payload, timeout=30)
                        if response.status_code == 200:
                            data = response.json()
                            all_products.extend(data.get("products", []))
                            
                            status_text.text(f"Descargados: {len(all_products)}/{total_products}")
                            progress_bar.progress(min(len(all_products) / total_products, 1.0))
                        else:
                            st.warning(f"⚠️ Error en lote (offset {offset}): {response.status_code}")
                            break
                        
                        offset += limit
                        time.sleep(0.3)
                    
                    progress_bar.empty()
                    status_text.empty()
                    
                    # ============================================================
                    # PASO 3: PROCESAR Y ANALIZAR
                    # ============================================================
                    wix_records = []
                    for product in all_products:
                        product_id = product.get('id')
                        sku = str(product.get('sku', ''))
                        visible = product.get('visible', False)
                        
                        price_data = product.get('price', {})
                        current_price = price_data.get('price', 0)
                        
                        stock_data = product.get('stock', {})
                        current_quantity = stock_data.get('quantity', 0)
                        
                        wix_records.append({
                            'product_id': product_id,
                            'SKU': sku,
                            'current_price': current_price,
                            'current_quantity': current_quantity,
                            'visible': visible
                        })
                    
                    wix_df = pd.DataFrame(wix_records)
                    
                    # Merge y análisis
                    merged_data = pd.merge(wix_df, data_ERP, on='SKU', how='inner')
                    merged_data['Precio'].fillna(0, inplace=True)
                    merged_data['Inventario'].fillna(0, inplace=True)
                    merged_data['should_be_visible'] = merged_data['Inventario'] > 0
                    
                    merged_data['needs_price_update'] = merged_data['current_price'] != merged_data['Precio']
                    merged_data['needs_inventory_update'] = merged_data['current_quantity'] != merged_data['Inventario']
                    merged_data['needs_visibility_update'] = merged_data['visible'] != merged_data['should_be_visible']
                    
                    merged_data['needs_update'] = (
                        merged_data['needs_price_update'] | 
                        merged_data['needs_inventory_update'] | 
                        merged_data['needs_visibility_update']
                    )
                    
                    to_update = merged_data[merged_data['needs_update'] == True]
                    
                    # ============================================================
                    # MOSTRAR RESULTADOS DEL ANÁLISIS
                    # ============================================================
                    st.success("✅ Análisis completado")
                    
                    st.markdown("---")
                    st.subheader("📊 Resumen del Análisis")
                    
                    # Métricas principales
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric(
                            label="🌐 Total en Wix",
                            value=total_products,
                            help="Total de productos en tu tienda Wix"
                        )
                    
                    with col2:
                        st.metric(
                            label="🔗 Encontrados en ERP",
                            value=len(merged_data),
                            help="Productos que existen tanto en Wix como en ERP"
                        )
                    
                    with col3:
                        st.metric(
                            label="✏️ A Modificar",
                            value=len(to_update),
                            help="Productos que requieren actualización"
                        )
                    
                    with col4:
                        porcentaje = (len(to_update) / len(merged_data) * 100) if len(merged_data) > 0 else 0
                        st.metric(
                            label="📈 % a Modificar",
                            value=f"{porcentaje:.1f}%"
                        )
                    
                    # Desglose detallado
                    if len(to_update) > 0:
                        st.markdown("#### 🔍 Desglose de Cambios")
                        
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            price_changes = to_update['needs_price_update'].sum()
                            st.metric("💰 Actualizarán Precio", price_changes)
                        
                        with col2:
                            inventory_changes = to_update['needs_inventory_update'].sum()
                            st.metric("📦 Actualizarán Inventario", inventory_changes)
                        
                        with col3:
                            visibility_changes = to_update['needs_visibility_update'].sum()
                            st.metric("👁️ Cambiarán Visibilidad", visibility_changes)
                        
                        # Preview de productos a modificar
                        st.markdown("#### 📋 Preview de Productos a Modificar (primeros 20)")
                        preview_df = to_update[['SKU', 'current_price', 'Precio', 'current_quantity', 'Inventario', 'visible', 'should_be_visible']].head(20)
                        preview_df.columns = ['SKU', 'Precio Actual', 'Precio Nuevo', 'Inv. Actual', 'Inv. Nuevo', 'Visible Actual', 'Visible Nuevo']
                        st.dataframe(preview_df, use_container_width=True)
                    else:
                        st.info("🎉 Todos los productos están actualizados. No se requieren cambios.")
                    
                except Exception as e:
                    st.error(f"❌ Error durante el análisis: {str(e)}")
                    st.code(str(e))
        
        # ====================================================================
        # SINCRONIZACIÓN (modificar productos)
        # ====================================================================
        if sync_button:
            
            with st.spinner('🌐 Conectando con Wix y descargando productos...'):
                try:
                    # ============================================================
                    # PASO 1: CARGAR DATOS DEL ERP
                    # ============================================================
                    st.info("📂 Paso 1/5: Cargando datos del ERP...")
                    
                    data_ERP = pd.read_csv(uploaded_file_erp, delimiter=';', encoding='latin1')
                    data_ERP = data_ERP[data_ERP['Codpro'].notna() & ~(data_ERP['Codpro'].isin(['', ' ']) | (data_ERP['Codpro'].str.contains('\x1a', na=False)))]
                    data_ERP = data_ERP[["Codpro", "Nompro", "Valuni", "us01", "us02"]]
                    data_ERP['us01'] = data_ERP['us01'].fillna(0)
                    data_ERP['us02'] = data_ERP['us02'].fillna(0)
                    data_ERP["Inventario_Bogota"] = data_ERP["us01"] + data_ERP["us02"]
                    data_ERP = data_ERP[["Codpro", "Valuni", "Inventario_Bogota"]]
                    data_ERP.rename(columns={'Codpro': 'SKU', 'Valuni': 'Precio', 'Inventario_Bogota': 'Inventario'}, inplace=True)
                    data_ERP['SKU'] = data_ERP['SKU'].astype(str)
                    
                    st.success(f"✅ ERP cargado: {len(data_ERP)} productos")
                    
                    # ============================================================
                    # PASO 2: DESCARGAR PRODUCTOS DE WIX
                    # ============================================================
                    st.info("🌐 Paso 2/5: Descargando productos de Wix...")
                    
                    all_products = []
                    offset = 0
                    limit = 100
                    url = f"{base_url}/v1/products/query"
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Primera llamada para obtener el total
                    payload = {
                        "includeHiddenProducts": True,
                        "query": {"paging": {"limit": limit, "offset": 0}}
                    }
                    
                    response = requests.post(url, headers=headers, json=payload, timeout=30)
                    
                    if response.status_code != 200:
                        st.error(f"❌ Error al conectar con Wix: {response.status_code}")
                        st.code(response.text)
                        st.stop()
                    
                    data = response.json()
                    total_products = data.get("totalResults", 0)
                    all_products.extend(data.get("products", []))
                    
                    status_text.text(f"Descargados: {len(all_products)}/{total_products}")
                    progress_bar.progress(min(len(all_products) / total_products, 1.0))
                    
                    # Continuar descargando el resto
                    offset = limit
                    while len(all_products) < total_products:
                        payload["query"]["paging"]["offset"] = offset
                        
                        response = requests.post(url, headers=headers, json=payload, timeout=30)
                        if response.status_code == 200:
                            data = response.json()
                            all_products.extend(data.get("products", []))
                            
                            status_text.text(f"Descargados: {len(all_products)}/{total_products}")
                            progress_bar.progress(min(len(all_products) / total_products, 1.0))
                        else:
                            st.warning(f"⚠️ Error en lote (offset {offset}): {response.status_code}")
                            break
                        
                        offset += limit
                        time.sleep(0.3)
                    
                    st.success(f"✅ Productos descargados: {len(all_products)}")
                    
                    # ============================================================
                    # PASO 3: PROCESAR DATOS DE WIX
                    # ============================================================
                    st.info("🔄 Paso 3/5: Procesando datos de Wix...")
                    
                    wix_records = []
                    for product in all_products:
                        product_id = product.get('id')
                        sku = str(product.get('sku', ''))
                        visible = product.get('visible', False)
                        
                        price_data = product.get('price', {})
                        current_price = price_data.get('price', 0)
                        
                        stock_data = product.get('stock', {})
                        current_quantity = stock_data.get('quantity', 0)
                        
                        wix_records.append({
                            'product_id': product_id,
                            'SKU': sku,
                            'current_price': current_price,
                            'current_quantity': current_quantity,
                            'visible': visible
                        })
                    
                    wix_df = pd.DataFrame(wix_records)
                    
                    # ============================================================
                    # PASO 4: HACER MERGE
                    # ============================================================
                    st.info("🔗 Paso 4/5: Combinando datos...")
                    
                    merged_data = pd.merge(wix_df, data_ERP, on='SKU', how='inner')
                    merged_data['Precio'].fillna(0, inplace=True)
                    merged_data['Inventario'].fillna(0, inplace=True)
                    merged_data['should_be_visible'] = merged_data['Inventario'] > 0
                    
                    merged_data['needs_price_update'] = merged_data['current_price'] != merged_data['Precio']
                    merged_data['needs_inventory_update'] = merged_data['current_quantity'] != merged_data['Inventario']
                    merged_data['needs_visibility_update'] = merged_data['visible'] != merged_data['should_be_visible']
                    
                    merged_data['needs_update'] = (
                        merged_data['needs_price_update'] | 
                        merged_data['needs_inventory_update'] | 
                        merged_data['needs_visibility_update']
                    )
                    
                    to_update = merged_data[merged_data['needs_update'] == True]
                    
                    st.success(f"✅ Productos que requieren actualización: {len(to_update)}")
                    
                    if len(to_update) == 0:
                        st.info("🎉 Todos los productos están actualizados. No se requieren cambios.")
                        st.stop()
                    
                    # Mostrar preview
                    st.subheader("📋 Preview de Cambios")
                    preview_df = to_update[['SKU', 'current_price', 'Precio', 'current_quantity', 'Inventario', 'visible', 'should_be_visible']].head(20)
                    st.dataframe(preview_df)
                    
                    # ============================================================
                    # PASO 5: ACTUALIZAR PRODUCTOS
                    # ============================================================
                    st.info("🚀 Paso 5/5: Actualizando productos en Wix...")
                    
                    progress_bar_update = st.progress(0)
                    status_text_update = st.empty()
                    
                    stats = {'success': 0, 'failed': 0, 'errors': []}
                    total_to_update = len(to_update)
                    
                    for idx, row in to_update.iterrows():
                        product_id = row['product_id']
                        sku = row['SKU']
                        new_price = float(row['Precio'])
                        new_quantity = int(row['Inventario'])
                        new_visible = bool(row['should_be_visible'])
                        
                        status_text_update.text(f"Actualizando {stats['success'] + stats['failed'] + 1}/{total_to_update}: SKU {sku}")
                        
                        # Obtener detalles del producto (revision y variant_id)
                        try:
                            details_url = f"{base_url}/v3/products/{product_id}"
                            details_response = requests.get(details_url, headers=headers, timeout=10)
                            
                            if details_response.status_code != 200:
                                stats['failed'] += 1
                                stats['errors'].append({'sku': sku, 'error': 'No se pudo obtener detalles'})
                                continue
                            
                            details_data = details_response.json()
                            product_details = details_data.get('product', {})
                            revision = product_details.get('revision', '1')
                            
                            variants_info = product_details.get('variantsInfo', {})
                            variants = variants_info.get('variants', [])
                            variant_id = variants[0].get('id') if variants else None
                            
                            if not variant_id:
                                stats['failed'] += 1
                                stats['errors'].append({'sku': sku, 'error': 'No se encontró variant_id'})
                                continue
                            
                            # Actualizar producto
                            update_url = f"{base_url}/v3/products-with-inventory/{product_id}"
                            update_payload = {
                                "product": {
                                    "id": product_id,
                                    "revision": revision,
                                    "visible": new_visible,
                                    "variantsInfo": {
                                        "variants": [{
                                            "id": variant_id,
                                            "choices": [],
                                            "price": {
                                                "actualPrice": {
                                                    "amount": str(new_price)
                                                }
                                            }
                                        }]
                                    }
                                }
                            }
                            
                            update_response = requests.patch(update_url, headers=headers, json=update_payload, timeout=10)
                            
                            if update_response.status_code == 200:
                                stats['success'] += 1
                            else:
                                stats['failed'] += 1
                                error_msg = update_response.json().get('message', 'Error desconocido')
                                stats['errors'].append({'sku': sku, 'error': error_msg})
                            
                            time.sleep(0.3)  # Rate limiting
                            
                        except Exception as e:
                            stats['failed'] += 1
                            stats['errors'].append({'sku': sku, 'error': str(e)})
                        
                        progress_bar_update.progress((stats['success'] + stats['failed']) / total_to_update)
                    
                    # ============================================================
                    # MOSTRAR RESULTADOS
                    # ============================================================
                    st.success("✅ ¡Sincronización completada!")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Procesados", total_to_update)
                    with col2:
                        st.metric("✅ Exitosos", stats['success'])
                    with col3:
                        st.metric("❌ Fallidos", stats['failed'])
                    
                    if stats['errors']:
                        st.error("🔴 Productos con errores:")
                        error_df = pd.DataFrame(stats['errors'])
                        st.dataframe(error_df)
                    
                    # Generar reporte descargable
                    report_df = to_update[['SKU', 'current_price', 'Precio', 'current_quantity', 'Inventario', 'visible', 'should_be_visible']]
                    report_csv = report_df.to_csv(index=False, encoding='utf-8-sig')
                    
                    st.download_button(
                        label="📥 Descargar Reporte Completo",
                        data=report_csv,
                        file_name=f"wix_sync_report_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
                    
                except Exception as e:
                    st.error(f"❌ Error durante la sincronización: {str(e)}")
                    st.code(str(e))

# --- APLICACIÓN PRINCIPAL (NAVEGACIÓN) ---
def main():
    # Mostrar logo en la barra lateral
    try:
        image = Image.open("logo_transparente.png")
        st.sidebar.image(image, use_container_width=True)
    except FileNotFoundError:
        st.sidebar.warning("Logo no encontrado.")

    st.sidebar.title("Menú de Navegación")
    st.sidebar.markdown("Selecciona la plataforma que deseas actualizar:")

    # Menú de selección en la barra lateral
    opciones = [
        "Mercado Libre - Medellín", 
        "Mercado Libre - Bogotá",
        "Falabella",
        "Rappi Av.19, Blv y Cll 74",
        "Rappi Bvista y Oviedo",
        "Wix (CSV)",
        "Wix (API)"
    ]
    opcion = st.sidebar.selectbox("Plataforma:", opciones)

    # Título principal de la aplicación
    st.title("🚀 MarketMaster")

    # Lógica para mostrar la página correcta según la selección
    if opcion == "Mercado Libre - Medellín":
        pagina_meli_medellin()
    elif opcion == "Mercado Libre - Bogotá":
        pagina_meli_bogota()
    elif opcion == "Falabella":
        pagina_falabella()
    elif opcion == "Rappi Av.19, Blv y Cll 74":
        pagina_rappi_av19_blv_cll74()
    elif opcion == "Rappi Bvista y Oviedo":
        pagina_rappi_bvista_oviedo()
    elif opcion == "Wix (CSV)":
        pagina_wix()
    elif opcion == "Wix (API)":
        pagina_wix_api()

    st.sidebar.info("Esta app centraliza la actualización de inventarios y precios en múltiples plataformas.")

if __name__ == "__main__":
    main()
